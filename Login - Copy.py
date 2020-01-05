#!/usr/bin/env python
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
from tkinter import *

#mesazh per te informuar 
mesazh=''

#info layer
class Excel():
    ex = load_workbook('sample.xlsx')
    names = ex['Sheet']
    
    aktiv = ex['aktivitet']
    admin = ex['admin']
    aa = admin['B2'].value
    na = admin['A2'].value 
    nameList = []

#mbledh listen
def findEmpty(ex):
    for ind,row in enumerate(ex.names):
        if row[0].value == row[1].value==row[2].value ==None:
            ex.na == ind
            break
    for ind,row in enumerate(ex.aktiv):
        if row[0].value == row[1].value==row[2].value ==None:
            ex.aa == ind
            break
    return ex

#degjon per ndryshime 
def mainLoop(ex):
    while True:
        i = input()
        if i !=False:
            aktivitet(ex,i)

#kerkon emrat 
def loadNames(ex):
    ex.nameList=[]
    for row in ex.names:
        step =[]
        for cell in row:
            step.append(cell.value)
        ex.nameList.append(step)
    return ex.nameList

#shton emrat ne liste
def addName(ex,name,title,salt,mesazh):
    #gjenero salt
    
    
    for row in ex.nameList:
        if row[0] == name:
            mesazh = 'Ky emer eshte ne liste'
            print(mesazh)
            return False
    else:
        ex.names['A'+ str(ex.na)]=name
        ex.names['B'+ str(ex.na)]=title
        ex.names['C'+ str(ex.na)]=salt
        ex.na= ex.na+1
        ex.admin['A2'].value = ex.na
        refresh(ex)

#rifreskon
def refresh(ex):
    ex = findEmpty(ex)
    ex.ex.save('sample.xlsx')
    ex.ex = load_workbook('sample.xlsx')
    
    ex.names = ex.ex['Sheet']
    
    ex.aktiv = ex.ex['aktivitet']
    ex.admin = ex.ex['admin']
    ex.aa = ex.admin['B2'].value
    ex.na = ex.admin['A2'].value 
    print('refreshed')

#shton aktivitet ne liste
def aktivitet(ex,salt):
    aR = str(ex.aa)
    for row in ex.nameList:
        if row[2] == salt:
            
            ex.aktiv['A'+aR].value=datetime.datetime.now()
            ex.aktiv['A'+aR].number_format = '[$-F400]h:mm:ss\ AM/PM'
            ex.aktiv['B'+aR].value=salt
            ex.aktiv['C'+aR].value=row[1]
            ex.aktiv['D'+aR].value=row[0]
            ex.aa= 1 + ex.aa
            ex.admin['B2'].value = ex.aa
    refresh(ex)

#layer pwr te shfaqur info ne ekran
class Row:
    def keep(self,name,title,time,tip):
        self.name = name
        self.title = title
        self.time = time
        self.tip = tip
        def show(self):
            return [self.name,self.title,self.time,self.tip]

#grafika kryesore
class Application(Frame):
    global win,select
    win = Frame()
    aktivityList=[]
    def __init__(self,master=None):
        Frame.__init__(self,master)
        self.grid(rowspan=14,columnspan=8)
        self.createWidget()
    def deleteRow(self):
        print(ind)
        print(ex.aktiv['A'+str(ind)])
        return False
        ex.aktiv['A'+str(ind)].value=None
        ex.aktiv['B'+str(ind)].value=None
        ex.aktiv['C'+str(ind)].value=None
        ex.aktiv['D'+str(ind)].value=None
        ex.aktiv['E'+str(ind)].value=None

    def showAktiv(self):
        aktivList=[]
        for ind,row in enumerate(ex.aktiv):
            step=[]
            step.append(ind)
            for ind2,cell in enumerate(row):
                if ind2 == 1:
                    continue
                else:
                    step.append(cell.value)
            aktivList.append(step)
        #aktivList.pop(0)
        #print(aktivList)
        self.numList=Label(self,text="Numri")
        self.numList.grid(row=3,column=4)
        
        self.nameList=Label(self,text="Emri")
        self.nameList.grid(row=3,column=0)
        self.pozList=Label(self,text="Pozicioni")
        self.pozList.grid(row=3,column=1)
        self.kohaList=Label(self,text="Ora")
        self.kohaList.grid(row=3,column=2)
        for ind,rows in enumerate(aktivList[::-1]):
            
            if ind >10:
                break
            #tregon nr rendor
            self.rownrList=Button(self,text=str(rows[0]))
            self.rownrList.grid(row=(4+ind),column=4)
            #tregon emrin
            self.rownaList=Button(self,text=str(rows[3]))
            self.rownaList.grid(row=(4+ind),column=0)
            #tregon pozicionin
            self.rowpozList=Button(self,text=str(rows[2]))
            self.rowpozList.grid(row=(4+ind),column=1)
            #tregon oren
            self.rowpozList=Button(self,text=str(rows[1]))
            self.rowpozList.grid(row=(4+ind),column=2)
            #butoni delete
            self.rowDelete=Button(self,text="Fshi",command=(self.deleteRow,rows[0]+1))
            self.rowDelete.grid(row=(4+ind),column=5)
        
    def createWidget(self):
        self.addName = Button(self,text='Add Name',command=self.addNameScreen)
        self.addName.grid(row=0,column=0)
        self.showA()
    #merr info nga kutia dhe shton emrin e ri
    #ka problem per momentin
    def nSMbrapa(self):
        self.nameField.grid_forget()
        self.pozField.grid_forget()

        self.showAktiv()

        self.addName = Button(self,text='Add Name',command=self.addNameScreen)
        self.addName.grid()
        self.enterName.grid_forget()
        self.mbrapa.grid_forget()
        
    def enterName(self):
        
        name = self.nameField.get()
        #print(name)
        #print('ketu')
        poz = self.pozField.get()
        if name == '' or poz == '':
            print('bosh')
            return False
        addName(ex,name,poz,name,mesazh)
        self.nameField.grid_forget()
        self.pozField.grid_forget()

        self.addName = Button(self,text='Add Name',command=self.addNameScreen)
        self.addName.grid()
        self.enterName.grid_forget()
        self.mbrapa.grid_forget()
        
    def addNameScreen(self):
        self.addName.grid_forget()
        self.nameField = Entry(self)
        self.nameField.grid()

        #shto puzicionin
        self.pozField = Entry(self)
        self.pozField.grid()

        self.enterName = Button(self,text='Aprovo te dhenat',command=self.enterName)
        self.enterName.grid()
        self.mbrapa = Button(self,text="Mbrapa",command=self.nSMbrapa)
        self.mbrapa.grid()

    def deleteEntry(self):
        print("delete")
    def show(self):
        print("delete")
    
    def showA(self):
        frame1=Frame()
        frame1.grid()

        b1 = Button(frame1,text="delete",command=self.deleteEntry)
        b1.grid(row=0,column=1)

        b2 = Button(frame1,text="show",command=self.show)
        b2.grid(row=0,column=2)

        scroll = Scrollbar(frame1,orient=VERTICAL)
        self.select = Listbox(frame1,yscrollcommand=scroll.set,height=10)


        scroll.config(command=self.select.yview)
        scroll.grid(row=2,column=11,rowspan=10)

        self.select.grid(row=2)
            

def setSelect(ex,app):
    aklist =[]
    for elem in ex.aktiv:
        step=[]
        for ind,cel in enumerate(elem):
            if ind ==0:
                continue
            step.append(cel.value)
        aklist.append(step)
    app.select.delete(0,END)
    for name in aklist:
        app.select.insert(END,name)



ex = Excel()
refresh(ex)
print(ex.aktiv['A9'].number_format)
i = Row()
i.keep('name',"title","time","tip")
#print(names['A1'].value)
#for row in names:
#    print(row)
ex.nameList = loadNames(ex)
app = Application()
app.master.title('Pjeter')
setSelect(ex,app)
#app.win.mainloop()
#teste

#print(ex.nameList)
#addName(ex,'ball','ing','033',mesazh)
#aktivitet(ex,'arjol')

